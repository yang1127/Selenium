# xlsx
import openpyxl


def read_excel(excel_url, sheet_name):
    """读取excel文件"""
    wb = openpyxl.load_workbook(excel_url)
    # 获取sheet对象
    sheet = wb[sheet_name]
    # 打印最大行数、列数
    # print(sheet.max_row, sheet.max_column)
    # 取第一行第一列  -- "序号"
    # print(sheet.cell(1, 1).value)
    # 循环行数
    all_list = []
    for row in range(2, sheet.max_row + 1):  # 从标题下一行遍历
        # 循环列数
        row_list = []
        for col in range(1, sheet.max_column):
            row_list.append(sheet.cell(row, col).value)
        # print(row_list)

        all_list.append(row_list)
        print(all_list)
    return all_list


'''
if __name__ == "__main__":
    read_excel("/Users/yangzhiqi/PycharmProjects/Selenium/Data/data.xlsx", "export")
'''
